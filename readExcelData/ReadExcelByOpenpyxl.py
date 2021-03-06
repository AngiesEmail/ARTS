#!/usr/bin/python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

import sys
reload(sys)
sys.setdefaultencoding('utf8')

class ReadExcel(object):
	"""docstring for ReadExcel"""
	def __init__(self):
		super(ReadExcel, self).__init__()

	def getConfigData(self,path,tag,index):
		config = self.readExcelData(path,tag,index)
		if config != None:
			configToDict = self.convertRowDict(config[0],config[1])
			return configToDict

	def readExcelData(self,path,tag,index):
		wb = load_workbook(path)
		sheetNames = wb.sheetnames
		for name in sheetNames:
			if not ("$" in name):
				sheet = wb[name]
				result = self.readSheetData(sheet,tag)
				if result != None:
					keyList = self.readExcelColKeyList(sheet,index)
					return [keyList,result]

	def readExcelColKeyList(self,sheet,index):
		rowsList = sheet.rows
		rowData = self.readRowDataByIndex(rowsList,index)
		return rowData
	
	def readSheetData(self,sheet,tag):
		rowsList = sheet.rows
		return self.readRowData(rowsList,tag)

	def readRowData(self,rowsList,tag):
		for row in rowsList:
			if (type(row[0].value) != 'unicode' and str(row[0].value) == tag) or row[0].value == tag:
				return row

	def readRowDataByIndex(self,rowsList,index):
		originIndex = 0
		for row in rowsList:
			originIndex = originIndex + 1
			if index == originIndex:
				return row

	def convertRowDict(self,keys,data):
		result = {}
		colNum = len(keys)
		for index in xrange(colNum):
			keyCell = keys[index]
			key = keyCell.value
			if key != None:
				valueCell = data[index]
				result[key] = valueCell.value
		return result
