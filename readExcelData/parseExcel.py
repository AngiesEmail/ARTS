#!/usr/bin/python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

import json

import sys
reload(sys)
sys.setdefaultencoding('utf8')

def readLevelInfo(path,levelId):
	return readExcelData(path,levelId)

def readExcelData(path,tag):
	wb = load_workbook(path)
	sheetNames = wb.sheetnames
	for name in sheetNames:
		if not ("$" in name):
			sheet = wb[name]
			result = readSheetData(sheet,tag)
			if result != None:
				keyList = readExcelColKeyList(sheet,3)
				return [keyList,result]


def readExcelColKeyList(sheet,index):
	rowsList = sheet.rows
	rowData = readRowDataByIndex(rowsList,index)
	return rowData
			
def readSheetData(sheet,tag):
	rowsList = sheet.rows
	return readRowData(rowsList,tag)

def readRowData(rowsList,tag):
	for row in rowsList:
		if (type(row[0].value) != 'unicode' and str(row[0].value) == tag) or row[0].value == tag:
			return row

def readRowDataByIndex(rowsList,index):
	originIndex = 0
	for row in rowsList:
		originIndex = originIndex + 1
		if index == originIndex:
			return row

def convertRowDict(keys,data):
	result = {}
	colNum = len(keys)
	for index in xrange(colNum):
		keyCell = keys[index]
		key = keyCell.value
		if key != None:
			valueCell = data[index]
			result[key] = valueCell.value
	return result

def writeDataToExcel(wb,ws,data):
	values = data.items()
	length = len(data)
	for i in xrange(1,length):
		ws.cell(1,i).value = values[i][0]
		if values[i][1] != None:
			ws.cell(2,i).value = values[i][1]
	wb.save("test.xlsx")

def writeData(wb,ws,bIndex,data):
	values = data.items()
	length = len(data)
	for i in xrange(1,length):
		ws.cell(bIndex,i).value = values[i][0]
		if values[i][1] != None:
			ws.cell(bIndex+1,i).value = values[i][1]
	wb.save("test.xlsx")

def writeValues(wb,ws,bIndex,data):
	values = data.items()
	length = len(data)
	for i in xrange(1,length):
		if values[i][1] != None:
			ws.cell(bIndex,i).value = values[i][1]
	wb.save("test.xlsx")

def readEnemyInfo(data,key,enemyPath):
	enemyDict = {}
	if data.has_key(key) and data[key] != None:
		enemy = readExcelData(enemyPath,data[key])
		enemyDict = convertRowDict(enemy[0],enemy[1])
	return enemyDict

def parseEnemySkills(enemyInfo,skillPath):
	keyList = ["skill0","skill1(skillId,rate)","skill2","skill3"]
	skillList = {}
	for key in keyList:
		skillId = enemyInfo.has_key(key) and enemyInfo[key]
		print "查询：%s" % skillId
		if skillId != None:
			skillInfo = readExcelData(skillPath,skillId)
			if skillInfo != None:
				skillDict = convertRowDict(skillInfo[0],skillInfo[1])
				skillList[key] = skillDict
			else:
				print "SkillInfo找不到%s" % skillId
	return skillList

def parseSkillInfo(skillList,effectPath):
	effectInfo = {}
	allValues = skillList.items()
	for key,value in allValues:
		tag = value["tag"]
		effectData = value.has_key("effects") and value["effects"]
		effectList = None
		if ";" in effectData:
			effectList = effectData.split(";")
		elif "," in effectData:
			effectList = effectData.split(",")
		else:
			effectList.append(effectData)
		if effectList != None:
			for value in effectList:
				effect = readExcelData(effectPath,value)
				effectConfig = convertRowDict(effect[0],effect[1])
				effectInfo[tag] = effectConfig

	return effectInfo
				


if __name__ == "__main__":

	wb = Workbook()
	ws = wb.active
	rowIndex = 1

	levelId = "10101"
	path = "Development/LevelInfo.xlsx"
	enemyPath = "Development/EnemyInfo.xlsx"
	skillPath = "Development/SkillEffect.xlsx"
	effectPath = "Development/FightingEffectConfig.xlsx"

	print "查询表：LevelInfo"
	result = readLevelInfo(path,levelId)
	data = convertRowDict(result[0],result[1])
	print "写入数据：LevelInfo"
	# writeData(wb,ws,rowIndex,data)

	print "查询表：EnemyInfo"
	enemyInfo = []
	skillEffect = {}
	effectConfig = {}
	for index in xrange(1,6):
		key = "enemy%d" % (1)
		enemyDict = readEnemyInfo(data,key,enemyPath)
		enemyInfo.append(enemyDict)

	rowIndex = 4
	print "写入数据：EnemyInfo"
	# writeData(wb,ws,rowIndex,enemyInfo[0])
	rowIndex = 6
	enemyNum = len(enemyInfo)
	# for index in xrange(1,enemyNum):
	# 	writeValues(wb,ws,rowIndex,enemyInfo[index])
	# 	rowIndex = rowIndex + 1

	print "查询表：SkillInfo"
	enemySkills = []
	effectConfigList = {}
	for index in xrange(0,enemyNum):
		tag = enemyInfo[index]["tag"]
		skillList = parseEnemySkills(enemyInfo[index],skillPath)
		effectConfigList[tag] = parseSkillInfo(skillList,effectPath)
		enemySkills.append(skillList)

	finalResult = {}
	finalResult["LevelInfo"] = data
	finalResult["EnemyInfo"] = enemyInfo
	finalResult["SkillEffect"] = enemySkills
	finalResult["FightingEffectConfig"] = effectConfigList

	jsonData = json.dumps(finalResult) 
	f = open("data.json", 'w')
	f.write(jsonData)
	f.close()

		


	

